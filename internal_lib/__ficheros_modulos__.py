from __myLIMS_wrappers__ import *
from win11toast import notify
import pandas as pd
from openpyxl import load_workbook
from zipfile import BadZipFile

#Insertar filtro automaticamente al inicio de la columna
def insertarFiltro(dirExcel, colnames):
  wb = load_workbook(dirExcel)
  ws = wb.active
  ws.auto_filter.ref = f"A1:{chr(64+len(colnames))}{ws.max_row}"
  wb.save(dirExcel)

#Obtener el último ID agregado a una columna ID de un archivo Excel
def ObtenerIDExcel(dirExcel, ncol, colnames):
  regExcel = AbrirXLSX(dirExcel,colnames=colnames)
  try:
    return int(regExcel.iloc[-1, ncol])
  except (IndexError,ValueError):
    return 0    

#Obtener el último ID agregado a una columna ID de un archivo Excel
def CambiarEstadoIDxlsx(dirExcel, ID, colnames, estado, ncol_ID=0, ncol_ESTADO=1, except_kill=False):
  regExcel = AbrirXLSX(dirExcel, colnames=colnames)
  try:
    col = regExcel.columns.to_list()

    ID_col = col[ncol_ID] 
    ESTADO_col = col[ncol_ESTADO]
    
    regExcel[ESTADO_col] = regExcel[ESTADO_col].astype(str)
    regExcel.loc[regExcel[ID_col] == ID, ESTADO_col] = estado

  except (IndexError,ValueError):
    if except_kill: raise ExcepcionArchivo("Error al cambiar el estado del ID, la cantidad de columnas no coincide")
    else: print("Error al cambiar el estado del ID, la cantidad de columnas no coincide")

  while True:
    try:
        regExcel.drop_duplicates().to_excel(dirExcel, index=False)
        insertarFiltro(dirExcel, colnames)
        break
    
    except PermissionError:
        if except_kill: raise ExcepcionArchivo("Error de permisos, se abrio el archivo con otra aplicación, favor cerrar y presionar enter para continuar")
        else: 
          notify("Error de permisos", "Se abrio el archivo con otra aplicación, favor cerrar y presionar enter para continuar")
          input("Error de permisos, se abrio el archivo con otra aplicación, favor cerrar y presionar enter para continuar")

def ComprobarExcelAbierto(dirExcel, colnames):
  try:
    regExcel = AbrirXLSX(dirExcel, colnames=colnames)
    regExcel.to_excel(dirExcel, index=False)
    return False
  except PermissionError:
    return True

def ReestablecerXLSX(dirExcel, colnames, filtro=False):
  while True:
    try:
        print("Limpiando archivo excel de muestras\n")
        os.remove(dirExcel)
        AbrirXLSX(dirExcel=dirExcel, colnames=colnames, except_create=True, except_alert=False, except_kill=False)
        if filtro: insertarFiltro(dirExcel, colnames)
        break

    except PermissionError:
        notify("Error de permisos", "Se abrio el archivo con otra aplicación, favor cerrar y presionar enter para continuar")
        input("Cierre el archivo excel para continuar\n")
        
    except FileNotFoundError:
        AbrirXLSX(dirExcel=dirExcel, colnames=colnames, except_create=True, except_alert=False, except_kill=False)
        break

def AbrirXLSX(dirExcel, colnames, except_kill=True, except_create=True, except_alert=True):
    try:
        regExcel = pd.read_excel(dirExcel, keep_default_na=False, index_col=None)

        if colnames != regExcel.columns.to_list() and except_alert:
            if except_kill: raise ExcepcionArchivo(f"La cantidad de columnas no coincide, los datos pueden quedar corridos\nSe esparaba: {colnames}\nSe lee: {regExcel.columns.to_list()}\n")
            else: 
              print(f"La cantidad de columnas no coincide, los datos pueden quedar corridos\nSe esparaba: {colnames}\nSe lee: {regExcel.columns.to_list()}\n")
              
        return regExcel
    
    except FileNotFoundError as e:
      if not except_create: raise ExcepcionArchivo("No se encontro el archivo")
      else:
        if except_alert: print("No se encontro el archivo, se creara uno nuevo...\n")

    except BadZipFile as e:
        print("El archivo dirExcel parece estar corrupto debido a un cierre inesperado al guardar\nVolviendo a crear el archivo...\n")
        os.remove(dirExcel)

    regExcel = pd.DataFrame(columns=[colnames])
    regExcel.columns = colnames
    regExcel.to_excel(dirExcel, index=False)
    insertarFiltro(dirExcel, colnames)
      
    return regExcel
    
def ListaAColumnaXLSX(dirExcel, valores_fila, colname, colnames, except_kill=False, except_create=True, funcion_print=print):
  regExcel = AbrirXLSX(dirExcel, colnames=colnames, except_kill=except_kill, except_create=except_create)
  start_idx = len(regExcel)
  funcion_print(f"Partiendo en {start_idx} y agregando {len(valores_fila)} valores")
  try:    
    for idx,val in enumerate(valores_fila):
      funcion_print(f"Agregando {val} en {start_idx+idx}")
      regExcel.loc[start_idx + idx] = {colname: val}
      
  except ValueError as e:

    if except_kill: 
      raise ExcepcionArchivo("La cantidad de columnas no coincide")
    else:
      funcion_print(f"Error al guardar el archivo, la cantidad e columnas no coincide {e}")
  
  while True:
      try:
          regExcel.drop_duplicates().to_excel(dirExcel, index=False)
          insertarFiltro(dirExcel, colnames)
          break
      
      except PermissionError:
          if except_kill: raise ExcepcionArchivo("Error de permisos, se abrio el archivo con otra aplicación, favor cerrar y presionar enter para continuar")
          else:
            notify("Error de permisos", "Se abrio el archivo con otra aplicación, favor cerrar y presionar enter para continuar")
            input("Error de permisos, se abrio el archivo con otra aplicación, favor cerrar y presionar enter para continuar")


def FilaAgregarXLSX(dirExcel, valores_fila, colnames, except_kill=False, except_create=True):

  regExcel = AbrirXLSX(dirExcel, colnames=colnames, except_kill=except_kill, except_create=except_create)

  try:    
    regExcel.loc[len(regExcel.index)] = valores_fila
  
  except ValueError as e:

    if except_kill: 
      raise ExcepcionArchivo("La cantidad de columnas no coincide")
    else:
      print(f"Error al guardar el archivo, la cantidad e columnas no coincide {e}")
  
  while True:
      try:
          regExcel.drop_duplicates().to_excel(dirExcel, index=False)
          insertarFiltro(dirExcel, colnames)
          break
      
      except PermissionError:
          if except_kill: raise ExcepcionArchivo("Error de permisos, se abrio el archivo con otra aplicación, favor cerrar y presionar enter para continuar")
          else:
            notify("Error de permisos", "Se abrio el archivo con otra aplicación, favor cerrar y presionar enter para continuar")
            input("Error de permisos, se abrio el archivo con otra aplicación, favor cerrar y presionar enter para continuar")



def AgregarMuestraXLSX(dirExcel, ID_Muestra, colnames, ID_controles=["-"], except_kill=False, except_create=False):
    
    regExcel = AbrirXLSX(dirExcel, colnames=colnames, except_kill=except_kill, except_create=except_create)
    try:
      if "ID CONTROL" in colnames:
        for idx, ID_control in enumerate(ID_controles):
          regExcel.loc[len(regExcel.index)+idx] = [str(ID_Muestra), str(ID_control)]
      else:
        regExcel.loc[len(regExcel.index)] = str(ID_Muestra)

    except ValueError as e:

      if except_kill: 
        raise ExcepcionArchivo("La cantidad de columnas no coincide")
      else:
        print(f"Error al guardar el archivo, la cantidad e columnas no coincide {e}")

    
    while True:
        try:
            regExcel.drop_duplicates().to_excel(dirExcel, index=False)
            insertarFiltro(dirExcel, colnames)
            break
        
        except PermissionError:
            if except_kill: raise ExcepcionArchivo("La cantidad de columnas no coincide")
            else:
              notify("Error de permisos", "Se abrio el archivo con otra aplicación, favor cerrar y presionar enter para continuar")  
              input("Error de permisos, se abrio el archivo con otra aplicación, favor cerrar y presionar enter para continuar")


def ListaMuestraXLSX(dirExcel, colname, colnames, except_kill=False, except_create=True, except_alert=False):

    regExcel = AbrirXLSX(dirExcel, colnames=colnames, except_kill=except_kill, except_create=except_create, except_alert=except_alert)
    
    try:
      # return [str(_) if type(_)!=list else _[0] for _ in regExcel[colname].values.tolist()]
      return [str(_).replace(" ","") for _ in regExcel[colname].values.tolist() if str(_).replace(" ","").isdigit()]
    
    except KeyError:
      if except_kill: raise ExcepcionArchivo(f"Error al obtener lista, la columna {colname} no existe")
      else: 
        print(f"Error al obtener lista, la columna {colname} no existe")
        return []

    